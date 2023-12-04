from os import path, mkdir
import re
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.workbook.workbook import Workbook


def check_token_format(token: str):
    pattern = r"^[0-9A-Z]{5}-[0-9A-Z]{5}-[0-9A-Z]{5}-[0-9A-Z]{5}-[0-9A-Z]{5}$"
    return bool(re.match(pattern, token))


def read_tokens(file_path: str):
    if not path.exists(file_path):
        raise Exception("File not found")
    wb: Workbook = load_workbook(file_path)
    sheet: Worksheet = wb.active
    tokens = [cell.value for column in sheet for cell in column if check_token_format(cell.value)]
    wb.close()
    return tokens


def adjusting_width(sheet: Worksheet):
    for column in sheet.columns:
        max_length = 0
        column = [cell for cell in column]
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)
        sheet.column_dimensions[column[0].column_letter].width = adjusted_width
    return sheet


def write_result(filename: str, token_dict: dict):
    wb: Workbook = Workbook()
    sheet: Worksheet = wb.active
    sheet.append(['Token', 'Result'])

    for token, result in token_dict.items():
        sheet.append([token, result])

    sheet = adjusting_width(sheet)

    processed_dir = "processed"
    if not path.exists(processed_dir):
        mkdir(processed_dir)
    processed_file_path = path.join(processed_dir, f'{filename}.xlsx')

    wb.save(processed_file_path)
    wb.close()
